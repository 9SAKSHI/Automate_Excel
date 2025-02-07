import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter

#--------------------------------------------------------- The Root Directory -------------------------------------------------------------------


directory_name = 'Employee_eFolder'

root_directory = r"C:\Users\SNR23\Desktop\HR_DATA\PIF_info_extraction\MIS_Employee Folder"



#-------------------------------------------------Traversing The Directory Structure------------------------------------------------------------

# Function to get Excel files
def get_excel_files(root_dir, target_month, target_date):
    excel_files = []
    calendar_order = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    # Check if the provided year exists in the root directory
    if not os.path.exists(root_dir):
        raise FileNotFoundError(f"The specified year directory '{root_dir}' does not exist.")
       
    
    # Ensure the root directory is indeed a directory
    if os.path.isdir(root_dir):
            
            month_list = os.listdir(root_dir)
            sorted_months = sorted(month_list, key=lambda x: calendar_order.index(x))
            

            # Iterate over sorted months
            for month_folder in sorted_months:
                month_path = os.path.join(root_dir, month_folder)
                
                if os.path.isdir(month_path):
                    # List date folders and convert to integers
                    date_list = os.listdir(month_path)
                    int_date_list = [int(date) for date in date_list]
                    
                    sorted_date_list = sorted(int_date_list)  # Sorting the list in ascending order
                    # print(f"Month {month_folder} - list of Days :", sorted_date_list)

                    # Iterate over dates in the current month
                    for date_folder in sorted_date_list:
                        # Process all dates if the current month is before the target month
                        if calendar_order.index(month_folder) < calendar_order.index(target_month):
                            date_path = os.path.join(month_path, str(date_folder))
                            # print(f"Processing date folder: {date_path}")

                            
                            if os.path.isdir(date_path):  
                                for name_folder in os.listdir(date_path):
                                    name_path= os.path.join(date_path, name_folder)
                                    # Check for Excel files in the folder
                                    if os.path.exists(name_path):
                                        for file in os.listdir(name_path):
                                    
                                            if file.endswith(".xlsx") or file.endswith(".xls"):  # Check for Excel files
                                                file_path = os.path.join(name_path, file)
                                                # print(f"Found Excel file: {file_path}")
                                                excel_files.append(file_path)

                        # Process only dates up to the target date if it's the target month
                        elif month_folder == target_month and date_folder <= target_date:
                            date_path = os.path.join(month_path, str(date_folder))
                            # print(f"Processing folder: {date_path}")
                            if os.path.isdir(date_path):  # Ensure the folder exists
                                for name_folder in os.listdir(date_path):
                                    name_path= os.path.join(date_path, name_folder)


                            # Check for Excel files in the folder
                                    if os.path.exists(name_path):  # Ensure the folder exists
                                        for file in os.listdir(name_path):
                                            if file.endswith(".xlsx") or file.endswith(".xls"):  # Check for Excel files
                                                file_path = os.path.join(name_path, file)
                                                # print(f"Found Excel file: {file_path}")
                                                excel_files.append(file_path)

                            # Stop processing after the target date
                            if date_folder == target_date:
                                break

                # Stop processing months after the target month
                if month_folder == target_month:
                    break

    return excel_files


#--------------------------------------------------------Input Parameters-----------------------------------------------------------------------------



full_path = os.path.join(root_directory, directory_name)

year = input("Enter the year: ")

# Set the root directory
root_dir = os.path.join(full_path, year)   
if not os.path.exists(root_dir):
    raise FileNotFoundError(f"The specified year directory does not exist.")


calendar_order = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

Target_month = input("Enter the Month till which you want to fetch the data [Jan, Feb, Mar, Apr, May, Jun, Jul, Aug, Sep, Oct, Nov, Dec] :")
Target_date = int(input("Enter the Date:"))






#----------------------------------------------------Get key Value--------------------------------------------------------------------------



    
def get_keys_value(df, result_dict): 
    for key in result_dict.keys():
        value_to_find = key
        row = df.isin([value_to_find]).any(axis=1).idxmax()
        
        
        col_name = df.isin([value_to_find]).any(axis=0).idxmax()
        col_index = df.columns.get_loc(col_name)
    
        if col_index + 1 < len(df.columns):
            value = df.iloc[row, col_index + 1] 
            result_dict[key] = value 
        else:
            result_dict[key] = None

    return result_dict



 #------------------------------------------------Initialise The Dictionaries----------------------------------------------------------------

def initialize_the_dicts():
    MIS_data_dict = {'Sr. No.': None,
                    'Employee ID': None,
                    'Trigram': None,
                    'Emp Name': None,
                    'Location':"Nan",
                    'Months': None,
                    'DOJ': None,
                    'Manager Name': None,
                    'Business Title': None,
                    'Brand/Group': None,
                    'Category': None,
                    'First name': None,
                    'Middle Name': None,
                    'Last Name': None,
                    'Date of Birth':None,
                    'Gender': None,
                    'Marital Status': None,
                    'Date of Marriage': None,
                    "Father's Full Name": None,
                    "Monther's Full Name": None,
                    'PAN': None,
                    'Aadhar Card Number': None,
                    'Name as per Aadhar Card': None,
                    'Permanent Address': None,
                    'City': None,
                    'State': None,
                    'Pincode': None,
                    'Personal Email ID': None,
                    'Contact Number': None,
                    "Spouse's Name": None
    }

    trigram_mail_dict = {'First name': None,
                        'Middle Name': None,
                        'Last Name': None,
                        'Date of Birth':None,
                        'Gender': None,
                        'Marital Status': None,
                        'Date of Marriage': None,
                        "Father's Full Name": None,
                        "Monther's Full Name": None,
                        'PAN': None,
                        'Aadhar Card Number': None,
                        'Name as per Aadhar Card': None,
                        'Remark':None,
                        'Permanent Address': None,
                        'City': None,
                        'State': None,
                        'Pincode': None,
                        'Personal Email ID': None,
                        'Contact Number': None,
                        'Location': None,
                        'Skill': None,
                        'Conversion (Yes/No)': None
    }

    IT_mail_dict = {'Trigram': None,
                    'Name': None,
                    'Location': None,
                    'Brand':None,
                    'Manager': None,
                    'Contact Number': None,
                    'Personal Email ID': None,
                    'Conversion (Yes/No)':None,
                    'Onboarding Formalities':None
    }
    return MIS_data_dict, trigram_mail_dict, IT_mail_dict





#----------------------------------------------------Extract the key values from the dataframe--------------------------------------------------------



excel_files_list=get_excel_files(root_dir, Target_month, Target_date)



#------------------------------------------------------------------------------------------------------------------------------------------------------



MIS_data = []
trigram_mail = []
IT_mail = []

i = 1
for file_path in excel_files_list:

    # print(i, file_path)
    MIS_data_dict, trigram_mail_dict, IT_mail_dict = initialize_the_dicts()
    df = pd.read_excel(file_path)

    df.dropna(how="all",inplace=True)
    df=df.dropna(how="all",axis=1)
    df.drop(df.columns[2:5], axis=1, inplace=True)
    df.drop(df.index[68:], inplace=True)
    df.reset_index(inplace=True, drop=True)
    df.iloc[64] = df.iloc[64].apply(lambda x: x.strip() if isinstance(x, str) else x)
    df.iloc[63] = df.iloc[63].apply(lambda x: x.strip() if isinstance(x, str) else x)
    

    MIS_data_dict = get_keys_value(df, MIS_data_dict)
    MIS_data_dict['Sr. No.'] = i
    
    
    MIS_data_dict['Aadhar Card Number'] = str(MIS_data_dict['Aadhar Card Number'])
    MIS_data.append(MIS_data_dict)
    MIS_data_dict['Location'] = None
    

    trigram_mail_dict = get_keys_value(df, trigram_mail_dict)
    trigram_mail_dict['Aadhar Card Number'] = str(trigram_mail_dict['Aadhar Card Number'])
    trigram_mail_dict['Location'] = None    
    trigram_mail.append(trigram_mail_dict)
   

    IT_mail_dict =  get_keys_value(df, IT_mail_dict)
    IT_mail.append(IT_mail_dict)
    i += 1
    first_name = trigram_mail_dict.get('First name', '')
    last_name = trigram_mail_dict.get('Last Name', '')
    IT_mail_dict['Name'] = f"{first_name} {last_name}".strip()
    IT_mail_dict['Location'] = None
    

    

MIS_data = pd.DataFrame(MIS_data)
MIS_data['Date of Birth'] = pd.to_datetime(MIS_data['Date of Birth']).dt.strftime('%d/%m/%Y')
MIS_data["Date of Marriage"] = pd.to_datetime(MIS_data['Date of Marriage']).dt.strftime('%d/%m/%Y')
trigram_mail = pd.DataFrame(trigram_mail)


trigram_mail['Date of Birth'] = pd.to_datetime(trigram_mail['Date of Birth']).dt.strftime('%d/%m/%Y')
trigram_mail['Date of Marriage'] = pd.to_datetime(trigram_mail['Date of Marriage']).dt.strftime('%d/%m/%Y')
IT_mail =  pd.DataFrame(IT_mail)




#-------------------------------------------------Save a data in excel sheets --------------------------------------------------


save_directory_Name= 'MIS' 

save_driectory_path=os.path.join(root_directory, save_directory_Name)
excel_file = os.path.join(save_driectory_path, f'MIS_CY_{year}.xlsx')



#-------------------------------------------------DataFrame to Excel Sheet-----------------------------------------------------


with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
    MIS_data.to_excel(writer, sheet_name='MIS Data', index=False)
    trigram_mail.to_excel(writer, sheet_name='Trigram mail format', index=False)
    IT_mail.to_excel(writer, sheet_name='IT mail format', index=False)




#-------------------------------------------------Style the excel sheets -----------------------------------------------------

wb = load_workbook(excel_file)
ws = wb.active

dark_blue_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") 
light_blue_fill = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid") 
header_font = Font(bold=True, color="FFFFFF") 
border_style = Border(
    left=Side(border_style="thin", color="000000"),
    right=Side(border_style="thin", color="000000"),
    top=Side(border_style="thin", color="000000"),
    bottom=Side(border_style="thin", color="000000")
)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]  
    columns_list = list(ws.columns)

    if sheet_name == 'Trigram mail format':
        total_columns = len(list(ws.columns))  
        for col_num, cell in enumerate(ws[1], start=1):  
            if col_num > total_columns - 2:  
                cell.fill = dark_blue_fill  
            else:
                cell.fill = light_blue_fill  
            cell.font = header_font  
            cell.border = border_style  
    elif sheet_name == "MIS Data":
        for col_num, cell in enumerate(ws[1], start=1): 
            if col_num <= 11:
                cell.fill = dark_blue_fill
                
            else:
                cell.fill = light_blue_fill  
            cell.font = header_font 
            cell.border = border_style 
             
    else:
        for cell in ws[1]:  
            cell.fill = light_blue_fill
            cell.font = header_font
            cell.border = border_style

    # Adjust column widths
    for col_num, column in enumerate(columns_list, start=1):
        max_length = 0
        column = get_column_letter(col_num)  
        
        # Find the maximum length of the column name and data
        for row in ws[column]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        
        adjusted_width = (max_length + 4)
        ws.column_dimensions[column].width = adjusted_width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(columns_list)):
        for cell in row:
            cell.border = border_style 

wb.save(excel_file)



#******************************************************************************************************************************************************************